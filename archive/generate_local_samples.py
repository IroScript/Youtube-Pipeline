import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = r"C:\Users\Irak\Desktop\deskTop\Cash in Hand and Dic Adjustment"
OUTPUT_DIR = os.path.join(BASE_DIR, "Depot_Wise_Local_Samples")

# Style Config
FONT_NAME = "Segoe UI"
COLOR_HEADER_FILL = "1E293B"      # Dark Slate
COLOR_HEADER_TEXT = "FFFFFF"      # White
COLOR_SUBHEADER_FILL = "F1F5F9"   # Very Light Gray
COLOR_ZEBRA_A = "FFFFFF"
COLOR_ZEBRA_B = "F8FAFC"
COLOR_MINT_FILL = "ECFDF5"        # Light Mint for totals
COLOR_MINT_TEXT = "065F46"        # Dark Green for totals

font_title = Font(name=FONT_NAME, size=16, bold=True, color="1E293B")
font_header = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_HEADER_TEXT)
font_bold = Font(name=FONT_NAME, size=10, bold=True)
font_regular = Font(name=FONT_NAME, size=10, bold=False)
font_total = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_MINT_TEXT)

fill_header = PatternFill("solid", fgColor=COLOR_HEADER_FILL)
fill_subheader = PatternFill("solid", fgColor=COLOR_SUBHEADER_FILL)
fill_zebra_a = PatternFill("solid", fgColor=COLOR_ZEBRA_A)
fill_zebra_b = PatternFill("solid", fgColor=COLOR_ZEBRA_B)
fill_mint = PatternFill("solid", fgColor=COLOR_MINT_FILL)

bd_thin = Side(style="thin", color="CBD5E1")
border_cell = Border(left=bd_thin, right=bd_thin, top=bd_thin, bottom=bd_thin)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

def format_sheet(ws, title_text, start_row, num_cols):
    ws.views.sheetView[0].showGridLines = True
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title_text
    title_cell.font = font_title
    title_cell.alignment = align_left
    ws.row_dimensions[1].height = 35
    
    # Sub-title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Generated on 2026-07-12 | Target Month: JUL'26 | Status: Draft Local Sample"
    subtitle_cell.font = Font(name=FONT_NAME, size=10, italic=True)
    subtitle_cell.alignment = align_left
    ws.row_dimensions[2].height = 20

def save_workbook_safely(wb, filename):
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"Saved: {path}")

def generate_reports():
    raw_data_path = os.path.join(BASE_DIR, "raw_compiled_data.json")
    if not os.path.exists(raw_data_path):
        print(f"Error: {raw_data_path} not found. Run compile_raw_data.py first.")
        return
        
    with open(raw_data_path, "r") as f:
        data = json.load(f)
        
    dates_desc = data['dates_descending']
    # Reverse to get chronological order (1st to 31st)
    dates_asc = list(reversed(dates_desc))
    records = data['records']
    
    # Group records by Depot
    depot_groups = {}
    for r in records:
        d = r['depot']
        if d not in depot_groups:
            depot_groups[d] = []
        depot_groups[d].append(r)
        
    # Generate Depot-Wise files
    for depot_name, depot_recs in depot_groups.items():
        wb = openpyxl.Workbook()
        
        # ─── TAB 1: Person-Wise Detail ───
        ws_det = wb.active
        ws_det.title = "Person_Wise_Detail"
        
        headers_det = ["Zone", "FM Name", "Market Name", "MPO Code", "Name/Position", "Designation", "Is Vacant"] + dates_asc + ["Total"]
        format_sheet(ws_det, f"{depot_name} DEPOT - Cash in Hand Detail (Person-Wise)", 4, len(headers_det))
        
        # Write headers at row 4
        ws_det.row_dimensions[4].height = 26
        for c_idx, h in enumerate(headers_det, 1):
            cell = ws_det.cell(row=4, column=c_idx)
            cell.value = h
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_cell
            
        # Freeze panes so headers and row metadata remain visible
        ws_det.freeze_panes = "H5"
        
        # Write records
        row_idx = 5
        for r in sorted(depot_recs, key=lambda x: (x['zone'], x['fm_name'], x['market_name'], x['mpo_code'])):
            ws_det.row_dimensions[row_idx].height = 20
            fill_row = fill_zebra_b if row_idx % 2 == 0 else fill_zebra_a
            
            # Metadata columns
            metadata = [
                r['zone'],
                r['fm_name'],
                r['market_name'],
                r['mpo_code'],
                r['person_name'] if r['person_name'] else r['da_name'],
                r['designation'],
                "Y" if r['is_vacant'] else ""
            ]
            
            for c_idx, val in enumerate(metadata, 1):
                cell = ws_det.cell(row=row_idx, column=c_idx)
                cell.value = val
                cell.font = font_regular
                cell.fill = fill_row
                cell.border = border_cell
                cell.alignment = align_left if c_idx in (2, 3, 5) else align_center
                
            # Date values
            col_idx = 8
            for date_str in dates_asc:
                val = r['daily_values'].get(date_str, 0)
                cell = ws_det.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = font_regular
                cell.fill = fill_row
                cell.border = border_cell
                cell.alignment = align_right
                cell.number_format = '#,##0'
                col_idx += 1
                
            # Total Column Formula
            cell_tot = ws_det.cell(row=row_idx, column=col_idx)
            start_col = get_column_letter(8)
            end_col = get_column_letter(col_idx - 1)
            cell_tot.value = f"=SUM({start_col}{row_idx}:{end_col}{row_idx})"
            cell_tot.font = font_total
            cell_tot.fill = fill_mint
            cell_tot.border = border_cell
            cell_tot.alignment = align_right
            cell_tot.number_format = '#,##0'
            
            row_idx += 1
            
        # Auto-adjust column widths for TAB 1
        for col in ws_det.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_det.column_dimensions[col_letter].width = max(max_len + 3, 10)
        # Freeze dates column widths
        for day_c in range(8, len(headers_det)):
            ws_det.column_dimensions[get_column_letter(day_c)].width = 11
            
        # ─── TAB 2: Market-Wise Summary (Aggregated) ───
        ws_sum = wb.create_sheet("Market_Wise_Summary")
        headers_sum = ["Zone", "FM Name", "Market Name", "MPO Code"] + dates_asc + ["Total"]
        format_sheet(ws_sum, f"{depot_name} DEPOT - Cash in Hand Summary (Market-Wise)", 4, len(headers_sum))
        
        ws_sum.row_dimensions[4].height = 26
        for c_idx, h in enumerate(headers_sum, 1):
            cell = ws_sum.cell(row=4, column=c_idx)
            cell.value = h
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_cell
            
        ws_sum.freeze_panes = "E5"
        
        # Aggregate by (Zone, FM Name, Market Name, MPO Code)
        agg_map = {}
        for r in depot_recs:
            key = (r['zone'], r['fm_name'], r['market_name'], r['mpo_code'])
            if key not in agg_map:
                agg_map[key] = {date_str: 0 for date_str in dates_asc}
            for date_str in dates_asc:
                agg_map[key][date_str] += r['daily_values'].get(date_str, 0)
                
        row_idx = 5
        for key in sorted(agg_map.keys()):
            zone, fm, mkt, mpo_c = key
            ws_sum.row_dimensions[row_idx].height = 20
            fill_row = fill_zebra_b if row_idx % 2 == 0 else fill_zebra_a
            
            # Metadata
            metadata = [zone, fm, mkt, mpo_c]
            for c_idx, val in enumerate(metadata, 1):
                cell = ws_sum.cell(row=row_idx, column=c_idx)
                cell.value = val
                cell.font = font_regular
                cell.fill = fill_row
                cell.border = border_cell
                cell.alignment = align_left if c_idx in (2, 3) else align_center
                
            # Dates
            col_idx = 5
            for date_str in dates_asc:
                val = agg_map[key][date_str]
                cell = ws_sum.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = font_regular
                cell.fill = fill_row
                cell.border = border_cell
                cell.alignment = align_right
                cell.number_format = '#,##0'
                col_idx += 1
                
            # Total Formula
            cell_tot = ws_sum.cell(row=row_idx, column=col_idx)
            start_col = get_column_letter(5)
            end_col = get_column_letter(col_idx - 1)
            cell_tot.value = f"=SUM({start_col}{row_idx}:{end_col}{row_idx})"
            cell_tot.font = font_total
            cell_tot.fill = fill_mint
            cell_tot.border = border_cell
            cell_tot.alignment = align_right
            cell_tot.number_format = '#,##0'
            
            row_idx += 1
            
        # Adjust column widths for TAB 2
        for col in ws_sum.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_sum.column_dimensions[col_letter].width = max(max_len + 3, 10)
        for day_c in range(5, len(headers_sum)):
            ws_sum.column_dimensions[get_column_letter(day_c)].width = 11
            
        save_workbook_safely(wb, f"{depot_name}_DEPOT_CASH_IN_HAND_Summary.xlsx")
        
    # ─── NATIONAL SUMMARY FILE ───
    wb_nat = openpyxl.Workbook()
    
    # TAB 1: National Market Summary
    ws_nat_sum = wb_nat.active
    ws_nat_sum.title = "National_Market_Summary"
    headers_nat = ["Depot", "Zone", "FM Name", "Market Name", "MPO Code"] + dates_asc + ["Total"]
    format_sheet(ws_nat_sum, "NATIONAL - Cash in Hand Summary (Market-Wise)", 4, len(headers_nat))
    
    ws_nat_sum.row_dimensions[4].height = 26
    for c_idx, h in enumerate(headers_nat, 1):
        cell = ws_nat_sum.cell(row=4, column=c_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    ws_nat_sum.freeze_panes = "F5"
    
    # Aggregate National by (Depot, Zone, FM Name, Market Name, MPO Code)
    nat_agg = {}
    for r in records:
        key = (r['depot'], r['zone'], r['fm_name'], r['market_name'], r['mpo_code'])
        if key not in nat_agg:
            nat_agg[key] = {date_str: 0 for date_str in dates_asc}
        for date_str in dates_asc:
            nat_agg[key][date_str] += r['daily_values'].get(date_str, 0)
            
    row_idx = 5
    for key in sorted(nat_agg.keys(), key=lambda x: (x[0], x[1], x[2], x[3])):
        depot, zone, fm, mkt, mpo_c = key
        ws_nat_sum.row_dimensions[row_idx].height = 20
        fill_row = fill_zebra_b if row_idx % 2 == 0 else fill_zebra_a
        
        # Metadata
        metadata = [depot, zone, fm, mkt, mpo_c]
        for c_idx, val in enumerate(metadata, 1):
            cell = ws_nat_sum.cell(row=row_idx, column=c_idx)
            cell.value = val
            cell.font = font_regular
            cell.fill = fill_row
            cell.border = border_cell
            cell.alignment = align_left if c_idx in (1, 3, 4) else align_center
            
        # Dates
        col_idx = 6
        for date_str in dates_asc:
            val = nat_agg[key][date_str]
            cell = ws_nat_sum.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = font_regular
            cell.fill = fill_row
            cell.border = border_cell
            cell.alignment = align_right
            cell.number_format = '#,##0'
            col_idx += 1
            
        # Total Formula
        cell_tot = ws_nat_sum.cell(row=row_idx, column=col_idx)
        start_col = get_column_letter(6)
        end_col = get_column_letter(col_idx - 1)
        cell_tot.value = f"=SUM({start_col}{row_idx}:{end_col}{row_idx})"
        cell_tot.font = font_total
        cell_tot.fill = fill_mint
        cell_tot.border = border_cell
        cell_tot.alignment = align_right
        cell_tot.number_format = '#,##0'
        
        row_idx += 1
        
    for col in ws_nat_sum.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_nat_sum.column_dimensions[col_letter].width = max(max_len + 3, 10)
    for day_c in range(6, len(headers_nat)):
        ws_nat_sum.column_dimensions[get_column_letter(day_c)].width = 11
        
    save_workbook_safely(wb_nat, "NATIONAL_CASH_IN_HAND_Summary.xlsx")

if __name__ == "__main__":
    generate_reports()
