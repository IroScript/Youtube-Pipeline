import openpyxl

file_path = r"Depot_Wise_Local_Samples\NATIONAL_CASH_IN_HAND_Summary.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("Sheetnames:", wb.sheetnames)
ws_nat = wb["National_Market_Summary"]
print("\n--- National_Market_Summary ---")
print("Dimensions:", ws_nat.dimensions)
for r in range(4, 9):
    row_vals = [ws_nat.cell(row=r, column=c).value for c in range(1, 15)]
    print(f"Row {r}: {row_vals}")
